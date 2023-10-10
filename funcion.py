import time
import var
import action
import openpyxl
import logging
import checkdata
import retry
from retry import retry







logging.basicConfig(handlers=[logging.FileHandler(filename="C:/Users/Admin/PycharmProjects/pythonProject/emso.log",
                                                 encoding='utf-8', mode='a+')],
                    format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                    datefmt="%F %A %T",
                    level=logging.INFO)



def getRowCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)

def getColumnCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value
def writeData(file,sheetName,rowum,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowum,column=columnno).value = data
    wordbook.save(file)

rows = getRowCount(var.path_baocao,'Sheet1')




@retry(tries=3, delay=2, backoff=1, jitter=5, )
class login():
    def thanh_cong_tk_emso(self):
        action.login.login3(self, "s1456a2@gmail.com", "Neko101213111999")
        writeData(var.path_baocao, "Sheet1", 11, 2, "x")
        writeData(var.path_baocao, "Sheet1", 25, 2, "x")
        action.logout()
    def khong_thanh_cong_tk_emso1(self):
        action.login.login1(self, "truongv22@gmail.com", "Neko101213111999")
        time.sleep(2)
        writeData(var.path_baocao, "Sheet1", 12, 2, "x")
        writeData(var.path_baocao, "Sheet1", 28, 2, "x")
    def khong_thanh_cong_tk_emso2(self):
        time.sleep(2)
        action.login.login2(self, "","")
        writeData(var.path_baocao, "Sheet1", 13, 2, "x")
        writeData(var.path_baocao, "Sheet1", 29, 2, "x")
        time.sleep(1)
    def thanh_cong_tk_google1(self):
        action.login.login4(self, "testeremso@gmail.com", "Neko10121311")
        writeData(var.path_baocao, "Sheet1", 15, 2, "x")
        action.logout()

    def thanh_cong_tk_google2(self):
        action.login.login_google(self, "testeremso2@gmail.com", "Neko10121311")
        # action.logout()
    def chon_tk_dang_nhap_gan_day_chua_luu_mk(self):
        action.login.login_chon_tk_dang_nhap_gan_day_chua_luu_mk(self, "Neko101213111999")
        writeData(var.path_baocao, "Sheet1", 21, 2, "x")
        writeData(var.path_baocao, "Sheet1", 24, 2, "x")
        action.logout()
    def chon_tk_dang_nhap_gan_day_da_luu_mk(self):
        action.login.login_chon_tk_dang_nhap_gan_day_da_luu_mk(self)
        writeData(var.path_baocao, "Sheet1", 20, 2, "x")
        writeData(var.path_baocao, "Sheet1", 23, 2, "x")
        action.logout()
    def chon_tk_dang_nhap_gan_day_nhap_sai_pass(self):
        action.login.login_chon_tk_dang_nhap_gan_day_nhap_sai_pass(self, "truongvccds")
        writeData(var.path_baocao, "Sheet1", 18, 2, "x")
        writeData(var.path_baocao, "Sheet1", 26, 2, "x")
    def chon_tk_dang_nhap_gan_day_khong_nhap_pass(self):
        action.login.login_chon_tk_dang_nhap_gan_day_khong_nhap_pass(self)
        writeData(var.path_baocao, "Sheet1", 19, 2, "x")
        writeData(var.path_baocao, "Sheet1", 27, 2, "x")



@retry(tries=3, delay=2, backoff=1, jitter=5, )
def thongtincanhan_anhdaidien(self):
    action.login.login3(self, "truongvck33@gmail.com", "atgmj123456")
    action.anhdaidien.anhdaidien_themmoi(self)
    action.checkdata_be.trangcanhan_thongtincanhan_anhdaidien_themmoi(self)

    action.anhdaidien.anhdaidien_khac(self)
    action.anhdaidien.anhdaidien_tuanhcosan(self)
    action.anhdaidien.anhdaidien_themkhung(self)
    action.checkdata_be.trangcanhan_thongtincanhan_anhdaidien_themkhung(self)

@retry(tries=3, delay=2, backoff=1, jitter=5, )
def thongtincanhan_anhbia(self):
    action.login.login3(self, "truongvck33@gmail.com", "atgmj123456")
    action.anhbia.anhbia_tailen(self)
    action.checkdata_be.trangcanhan_thongtincanhan_anhbia_tailen(self)
    action.anhbia.anhbia_thayanh(self)
    action.anhbia.anhbia_chonanh(self)
    action.anhbia.anhbia_chinhsuavitri(self)



class gioithieu():
    def gioithieu_tongquan(self):
        action.login.login3(self,"truongvck33@gmail.com", "atgmj123456")

        action.tongquan.tongquan(self)
        action.checkdata_be.trangcanhan_gioithieu_tongquan(self)

        # action.tongquan.tongquan_dulieusai(self)
    def congviec_vahocvan(self):
        # action.login.login3(self, "truongvck33@gmail.com", "atgmj123456")
        action.cong_viec_va_hoc_van.congviecvahocvan_congviec(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_congviec(self)

        action.cong_viec_va_hoc_van.congviecvahocvan_daihoc(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_daihoc(self)

        action.cong_viec_va_hoc_van.congviecvahocvan_trunghoc(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_trunghoc(self)
    def xemsukientrongdoi_congviecvahocvan(self):
        # action.login.login3(self, "truongvck33@gmail.com", "atgmj123456")

        action.xemkientrongdoi.congviecvahocvan_congviec(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_congviec(self)
        #
        action.xemkientrongdoi.congviecvahocvan_daihoc(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_daihoc(self)

        action.xemkientrongdoi.congviecvahocvan_trunghoc(self)
        action.checkdata_be.trangcanhan_gioithieu_cvvahocvan_xemsukientrongdoi_trunghoc(self)
    def trangcanhan_gioithieu_congviecvahocvan_addthem(self):
        # action.login.login3(self, "truongvck33@gmail.com", "atgmj123456")
        action.cong_viec_va_hoc_van.congviecvahocvan_addthem(self)

    # @retry(tries=3, delay=2, backoff=1, jitter=5, )
    def trangcanhan_gioithieu_thongtincoban(self):
        action.login.login3(self, "truongvck33@gmail.com", "atgmj123456")
        action.thongtincoban.thongtincoban(self)
        action.checkdata_be.trangcanhan_gioithieu_thongtincoban(self)

        action.thongtincoban.thongtincoban_dulieusai(self)



